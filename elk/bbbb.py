#!/usr/bin/env python
# -*- coding:utf-8 -*-
# author: Zhoulishan
# email : zhou_mfk@163.com
# date  : 17/04/28 17:10:47
import sys
import time
import argparse
from es import GetData
from openpyxl import Workbook

def get_aggs(field=None, size=None):
    aggs = {}
    if size is None and field in ['request_time', 'bytes']:
        aggs = {
            "1": {
                "min": {
                    "field": field
                }
            },
            "2": {
                "max": {
                    "field": field
                }
            },
            "3": {
                "avg": {
                    "field": field
                }
            }
        }
    elif size is not None and field in ['clientip', 'response', 'request']:
        aggs = {
            "1": {
                "terms": {
                    "field": field,
                    "size": size,
                    "order": {
                        "_count": "desc"
                    }
                }
            }
        }
    elif field is None and size is None:
        aggs = {}
    else:
        aggs = {}
    return aggs


def an_p(data, total):
    if not data['timed_out']:
        temp_d1 = {}
        temp_d2 = {}
        if len(data['aggregations'].keys()) == 1:
            if data['aggregations']['1']['buckets']:
                temp_dict = {}
                temp_list = data['aggregations']['1']['buckets']
                for i in temp_list:
                    temp_dict[i['key']] = i['doc_count']
                if temp_dict:
                    for k in sorted(temp_dict.items(), lambda x, y: cmp(x[1], y[1]), reverse=True):
                        # print k[0], k[1], format(float(k[1]) / float(total), '.3%')
                        temp_d1[k[0]] = k[1]
                        temp_d2[k[0]] = format(float(k[1]) / float(total), '.3%')
        return temp_d1, temp_d2


def an_data(data, isok):
    if isok:
        if not data['timed_out']:
            d = {}
            if len(data['aggregations'].keys()) == 1:
                if data['aggregations']['1']['buckets']:
                    temp_dict = {}
                    temp_list = data['aggregations']['1']['buckets']
                    for i in temp_list:
                        temp_dict[i['key']] = i['doc_count']
                    if temp_dict:
                        for k in sorted(temp_dict.items(), lambda x, y: cmp(x[1], y[1]), reverse=True):
                            d[k[0]] = k[1]
            elif len(data['aggregations'].keys()) == 3:
                if data['aggregations']:
                    d['min'] = data['aggregations']['1']['value']
                    d['max'] = data['aggregations']['2']['value']
                    d['avg'] =  data['aggregations']['3']['value']
            return d
    else:
        if not data['timed_out']:
            print "查询到的总数据", data['hits']['total']


if __name__ == "__main__":
    wb = Workbook()
    p = 0
    t1 = "2017-05-15 00:00:00"
    t2 = "2017-06-15 23:59:59"
    with open('ng.txt') as f:
        fs = f.readlines()
        for i in fs:
            i = i.strip()
            t = i.split()

            q1 = "%s:%s" % (t[0], t[1])
            q2 = "%s:%s AND request_time:>3" % (t[0], t[1])
            L = ['clientip', 'response', 'request', 'request_time', 'bytes', 'total', 'gt_3']
            num = 0
            ws = wb.create_sheet(index=p, title=t[1])
            p = p + 1

            ws.merge_cells('A1:C1')
            #ws['A1'] = '数据采集时间【%s-%s】' % (t1.split()[0].replace('-',''), t2.split()[0].replace('-',''))
            ws['A1'] = '数据采集时间【%s-%s】' % (t1, t2)
            ws['A2'] = '应用名称'
            ws['B2'] = '域名'
            ws['C2'] = '总请求'

            for i in L:
                if i == 'total':
                    aggs = {}
                    b = GetData('graylog*', q1, aggs, t1, t2)
                    total = b.get_total()
                    ws['A3'] = '%s' % t[1]
                    ws['B3'] = '%s' % t[1]
                    ws['C3'] = total
                elif i == 'request_time':
                    aggs = get_aggs('request_time')
                    b = GetData('graylog*', q1, aggs, t1, t2)
                    d = an_data(b.get_data(), True)
                    ws.merge_cells('A5:C5')
                    ws['A5'] = '响应时间 s'
                    ws['A6'] = '最短'
                    ws['B6'] = '最长'
                    ws['C6'] = '平均'
                    ws['A7'] = d['min']
                    ws['B7'] = d['max']
                    ws['C7'] = d['avg']
                elif i == 'bytes':
                    aggs = get_aggs('bytes')
                    b = GetData('graylog*', q1, aggs, t1, t2)
                    d = an_data(b.get_data(), True)
                    ws.merge_cells('A9:C9')
                    ws['A9'] = '报文大小 bytes'
                    ws['A10'] = '最小'
                    ws['B10'] = '最大'
                    ws['C10'] = '平均'
                    ws['A11'] = d['min']
                    ws['B11'] = d['max']
                    ws['C11'] = d['avg']
                elif i == 'clientip':
                    aggs = get_aggs('clientip', 5)
                    b = GetData('graylog*', q1, aggs, t1, t2)
                    total = b.get_total()
                    d1, d2 = an_p(b.get_data(), total)
                    ws.merge_cells('A13:C13')
                    ws['A13'] = 'client ip top5'
                    ws['A14'] = '总数'
                    ws['B14'] = '占比'
                    ws['C14'] = 'IP'
                    for k in sorted(d1.items(), lambda x, y: cmp(x[1], y[1]), reverse=True):
                        ws.append([k[1], d2[k[0]], k[0]])

                elif i == 'response':
                    aggs = get_aggs('response', 30)
                    b = GetData('graylog*', q1, aggs, t1, t2)
                    total = b.get_total()
                    d1, d2 = an_p(b.get_data(), total)
                    ws.merge_cells('A22:C22')
                    ws['A22'] = '各状态码'
                    ws['A23'] = '状态码'
                    ws['B23'] = '数量'
                    ws['C23'] = '占比'
                    l = d1.keys()
                    num = len(l)
                    l.sort()
                    for i in l:
                        m = [i, d1[i], d2[i]]
                        ws.append(m)

                elif i == 'request':
                    aggs = get_aggs('request', 5)
                    b = GetData('graylog*', q2, aggs, t1, t2)
                    total = b.get_total()
                    d1, d2 = an_p(b.get_data(), total)
                    a = 26 + num
                    b = a + 1
                    ws.merge_cells('A%s:C%s' % (a, a))
                    ws['A'+ str(a)] = '大于3S'
                    ws['C'+ str(b)] = 'URL'
                    ws['A'+ str(b)] = '总数'
                    ws['B'+ str(b)] = '占比'
                    for k in sorted(d1.items(), lambda x, y: cmp(x[1], y[1]), reverse=True):
                        ws.append([k[1], d2[k[0]], k[0]])

    wb.save('report/%s~%s健康报告.xlsx' % (t1.split()[0], t2.split()[0]))
